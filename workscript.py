import os, openpyxl
import string

from openpyxl.workbook.workbook import Workbook
import requests, time,random
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import re
import urllib.parse

# from .bit_api import createBrowser, openBrowser, closeBrowser, deleteBrowser

ROOT = os.path.dirname(__file__) + os.sep

#############全局固定变量####################
ACCOUNT_FILES = ROOT + "accounts.xlsx"
OUTLOOK_URL = "https://www.microsoft.com/en-us/microsoft-365/outlook/email-and-calendar-software-microsoft-outlook-b?deeplink=%2Fowa%2F%3Frealm%3Doutlook.com&sdf=0"
LOOK_TARGET = "https://huaren.us/index.html?page=register.html"
PATH = r"D:\Program\Python3\chromedriver.exe"
GLOBAL_WAIT = 60
############################账号密码生成器#############

import requests
import json
import time

# 官方文档地址
# https://doc2.bitbrowser.cn/jiekou/ben-di-fu-wu-zhi-nan.html

# 此demo仅作为参考使用，以下使用的指纹参数仅是部分参数，完整参数请参考文档

url = "http://127.0.0.1:54345"
headers = {'Content-Type': 'application/json'}


def createBrowser():  # 创建或者更新窗口，指纹参数 browserFingerPrint 如没有特定需求，只需要指定下内核即可，如果需要更详细的参数，请参考文档
    json_data = {
        'name': 'google',  # 窗口名称
        'remark': '',  # 备注
        'proxyMethod': 2,  # 代理方式 2自定义 3 提取IP
        # 代理类型  ['noproxy', 'http', 'https', 'socks5', 'ssh']
        'proxyType': 'noproxy',
        'host': '',  # 代理主机
        'port': '',  # 代理端口
        'proxyUserName': '',  # 代理账号
        "browserFingerPrint": {  # 指纹对象
            'coreVersion': '112'  # 内核版本 112 | 104，建议使用112，注意，win7/win8/winserver 2012 已经不支持112内核了，无法打开
        }
    }

    res = requests.post(f"{url}/browser/update",
                        data=json.dumps(json_data), headers=headers).json()
    browserId = res['data']['id']
    print(browserId)
    return browserId


def updateBrowser():  # 更新窗口，支持批量更新和按需更新，ids 传入数组，单独更新只传一个id即可，只传入需要修改的字段即可，比如修改备注，具体字段请参考文档，browserFingerPrint指纹对象不修改，则无需传入
    json_data = {'ids': ['93672cf112a044f08b653cab691216f0'],
                 'remark': '我是一个备注', 'browserFingerPrint': {}}
    res = requests.post(f"{url}/browser/update/partial",
                        data=json.dumps(json_data), headers=headers).json()
    print(res)


def openBrowser(id):  # 直接指定ID打开窗口，也可以使用 createBrowser 方法返回的ID
    json_data = {"id": f'{id}'}
    res = requests.post(f"{url}/browser/open",
                        data=json.dumps(json_data), headers=headers).json()
    print(res)
    print(res['data']['http'])
    return res


def closeBrowser(id):  # 关闭窗口
    json_data = {'id': f'{id}'}
    requests.post(f"{url}/browser/close",
                  data=json.dumps(json_data), headers=headers).json()


def deleteBrowser(id):  # 删除窗口
    json_data = {'id': f'{id}'}
    print(requests.post(f"{url}/browser/delete",
          data=json.dumps(json_data), headers=headers).json())


#############################
class PwdUtils:

    @classmethod
    def makeAccount(cls,length=10):
        # 定义密码可能包含的字符集合：大小写字母和数字
        characters = string.ascii_letters + string.digits
        account = ''.join(random.choice(characters) for i in range(length))+str(1)
        return account


    @classmethod
    def makePwd(cls,length=10):
        # 定义密码可能包含的字符集合：大小写字母和数字
        characters = string.ascii_letters + string.digits
        account = ''.join(random.choice(characters) for i in range(length))+"a"+"1"
        return account+"@@"



##############################################
class ProxyWork:
    def __init__(self):
        self.session = requests.Session()

    def get_proxy(self):
        time.sleep(3)
        response = self.session.get(
            "http://api.proxy.ipidea.io/getBalanceProxyIp?num=1&return_type=txt&lb=1&sb=0&flow=1&regions=&protocol=http")
        data = response.text
        self.session.close()
        print("获取到代理:" + data)
        return data.split(":")[0], data.split(":")[1]


class EmailLogin:
    def __init__(self, ac, pd,driverPath,debuggerAddress):
        self.ac = ac
        self.pd = pd
        self.driverPath = driverPath
        self.debugAddress = debuggerAddress


    def click_if_exists(self, xpath, message, noWait=True, timeot=15):
        if noWait:
            self.driver.implicitly_wait(timeot)
        print(time.strftime("%Y-%m-%d %H:%M:%S",time.localtime(time.time()))+"操作:" + message)
        targets = self.driver.find_elements(By.XPATH, xpath)
        if targets:
            targets[0].click()
        self.driver.implicitly_wait(GLOBAL_WAIT)

    def click_it(self, xpath, message):
        print("操作:" + message)
        wait = WebDriverWait(self.driver, GLOBAL_WAIT, 0.5)
        ele = wait.until(EC.presence_of_element_located((By.XPATH, xpath)))
        ele.click()

    def login(self,account1,pwd1):
        chrome_options = webdriver.ChromeOptions()
        chrome_options.add_experimental_option("debuggerAddress", self.debugAddress)
        chrome_service = Service(self.driverPath)
        self.driver = webdriver.Chrome(service=chrome_service,options=chrome_options)
        self.driver.maximize_window()
        self.driver.implicitly_wait(GLOBAL_WAIT / 12)
        self.driver.get(OUTLOOK_URL)
        self.click_it('//*[contains(text(),"Sign in")]', "点击登录按钮")
        windows = self.driver.window_handles
        self.driver.switch_to.window(windows[-1])
        time.sleep(2)
        self.driver.find_element(By.XPATH, '//*[@id="i0116"]').send_keys(self.ac)
        self.click_it("//*[@id='idSIButton9']", "点击下一步")
        time.sleep(2)
        self.driver.find_element(By.XPATH, '//*[@id="i0118"]').send_keys(self.pd)
        time.sleep(2)
        self.driver.find_element(By.XPATH, "//*[@id='idSIButton9']").click()
        time.sleep(2)
        self.click_if_exists('//*[@id="iShowSkip"]', "点击跳过")
        self.click_if_exists('//*[@id="iNext"]', "点击下一步")
        self.click_if_exists('//*[@id="id__0"]', "查看是否又弹框")
        self.click_if_exists('//*[@id="pageContent"]/div/div/form/div[3]/div[1]/div/label/span', "点击不再显示对话框")
        self.click_if_exists("//*[@id='declineButton']", "点击否")
        self.click_if_exists('//*[@id="iCancel"]', "点击不谢谢")
        self.click_if_exists("//*[text()='确定']", "点击确定")
        time.sleep(5)
        for i in range(3):
            self.driver.refresh()
            time.sleep(2)
        url = self.driver.current_url
        if "office.com" in url:  # 如果进入的是Office
            self.click_if_exists('//*[@id="sre-dismiss-button"]/div/button/span/svg',"关闭X")
            self.driver.refresh()
            time.sleep(5)
            self.click_if_exists('//*[@id="tui-callout"]/div[3]/div/div/div[2]/div[2]/button/span', "关闭X")
            self.click_if_exists('//*[@id="Mail"]', "点击OUTLOOK")
            windows = self.driver.window_handles
            self.driver.switch_to.window(windows[-1])
            self.driver.find_element(By.XPATH, '//*[@id="i0116"]').send_keys(self.ac)
            self.click_it("//*[@id='idSIButton9']", "点击下一步")
        self.click_if_exists("//*[text()='OK']","关闭ok")
        self.click_it("//*[text()='Junk Email']", "垃圾箱")
        self.click_if_exists("//*[contains(text(),'<pwd-service-a@huaren.us>')]", "北美华人网")
        time.sleep(2)
        self.click_if_exists("//*[text()='Accept']", "Accept")
        self.driver.find_elements(By.XPATH, "//*[contains(text(),'安全注册链接')]")[0].click()
        time.sleep(2)
        self.click_if_exists("//*[text()='Accept']","Accept")
        target = self.driver.find_elements(By.XPATH, "//*[contains(text(),'Visit Link') or contains(text(),'安全注册链接')]") # 这里写的不够好，有更好的写法。貌似这个 返回的 target 和 原本要做的事情 对不上 了。

        print(self.driver.find_elements(By.XPATH, "//*[contains(text(),'安全注册链接')]")[0].text)

        target2 = self.driver.find_elements(By.XPATH, "//*[contains(text(),'30分钟后失效')]")
        print(target2[0].text)
        # 这些代码写的有点乱，各种调试，找 相关 element， 找url；后续需要优化下， 方法封装，变量命名； 之前写的那些element应该都失效了。
        text = target[0].text
        print(text)
        # 不通过 click 拿到 url 的方式
        url_pattern = r"验证码 \w+"
        print(re.findall(url_pattern, text)[0].split(" ")[1])
        url1 = "https://huaren.us/register.html?vcode=" + re.findall(url_pattern, text)[0].split(" ")[1]  # zhel
        print(url1)
        # encoded_url = urllib.parse.quote(url1)

        # self.driver.execute_script("window.open(" + url1 + ");")
        self.driver.execute_script(f"window.open(\"{url1}\");")

        # target[0].get_attribute("href"

        # self.driver.get(url)
        # if len(target)==1:
        #     target[0].click()
        # if len(target)==2:
        #     target[0].click()

        #切到最后一个窗口
        self.driver.refresh()
        windows = self.driver.window_handles[-1]
        self.driver.switch_to.window(windows)
        self.driver.find_elements(By.XPATH, '//*[@id="reg_username"]')[0].send_keys(account1)
        self.driver.find_elements(By.XPATH, '//*[@id="reg_username"]')[1].send_keys(pwd1)
        self.driver.find_elements(By.XPATH, '//*[@id="reg_username"]')[2].send_keys(pwd1)
        self.driver.find_element(By.XPATH, '//*[@id="root"]/div/div/div/form/div[5]/button').click()
        time.sleep(20)

class RegisterWork:
    def __init__(self, driverPath, debuggerAddress):
        self.driverPath = driverPath
        self.debugAddress=debuggerAddress
    def register(self, email):
        chrome_options = webdriver.ChromeOptions()
        chrome_options.page_load_strategy = 'eager'
        chrome_options.add_experimental_option("debuggerAddress", self.debugAddress)
        chrome_service = Service(self.driverPath)
        self.driver = webdriver.Chrome(service=chrome_service,options=chrome_options)
        self.driver.maximize_window()
        self.driver.get(LOOK_TARGET)
        # self.driver.implicitly_wait(GLOBAL_WAIT)
        # self.driver.implicitly_wait(10)
        time.sleep(5)
        # self.driver.find_element(By.XPATH, "//*[text()='同意']").click()
        print("点击同意")
        self.driver.find_element(By.XPATH, "//button[contains(@class, 'btn') and contains(@class, 'btn-blue') and text()='同意']").click()

        time.sleep(5)
        print("输入注册框")
        self.driver.find_element(By.XPATH, "//*[@class='login-form form-horizontal']//input").send_keys(email)
        time.sleep(5)
        self.driver.find_element(By.XPATH, "//*[text()='提交请求']").click()
        time.sleep(5)
        # bug fix: wait till the element is present
        WebDriverWait(self.driver, 60, 0.5).until(EC.presence_of_element_located((By.XPATH, "//p[contains(text(),'部分邮箱会将我们的邮件归类到Spam')]")))
        target = self.driver.find_element(By.XPATH, "//p[contains(text(),'部分邮箱会将我们的邮件归类到Spam')]")
        assert target != None
        time.sleep(5)


class WorkMng:
    def __init__(self):
        self.workbook = openpyxl.load_workbook(ACCOUNT_FILES)
        self.sheet = self.workbook.active

    def loadInfos(self):
        nrows = self.sheet.max_row
        total=[]
        for i in range(2, nrows + 1):
            print("row: " + str(i))
            mailAccount = self.sheet.cell(i, 1).value.strip()
            passwd = self.sheet.cell(i, 2).value.strip()
            userAccount = PwdUtils.makeAccount()
            userPwd = PwdUtils.makePwd()
            browser_id = createBrowser()
            try:
                print("账号:"+userAccount)
                print("密码:" + userPwd)
                res = openBrowser(browser_id)
                driverPath = res['data']['driver']
                debuggerAddress = res['data']['http']
                register = RegisterWork(driverPath, debuggerAddress)
                register.register(mailAccount)
                email = EmailLogin(mailAccount, passwd, driverPath, debuggerAddress)
                email.login(userAccount, userPwd)
            except Exception as e:
                import traceback
                traceback.print_exc()
                print("email:" + mailAccount + "注册失败")
            else:
                vliststr = "email:" + mailAccount + "注册成功"+"账号:"+userAccount+"密码:"+userPwd
                print(vliststr)
                self.sheet.cell(i, 3).value = userAccount
                self.sheet.cell(i, 4).value = userPwd
                self.workbook.save(ACCOUNT_FILES)
                total.append(vliststr)
                print("success excel write")
            finally:
                closeBrowser(browser_id)
                deleteBrowser(browser_id)
        print(total)

if __name__ == '__main__':
    work = WorkMng()
    work.loadInfos()
    # proxy = ProxyWork()
    # proxy.get_proxy()
    # print(PwdUtils.makeAccount())
    # print(PwdUtils.makePwd())
